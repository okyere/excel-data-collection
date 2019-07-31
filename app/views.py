# -*- coding: utf-8 -*-

from functools import wraps
from flask import render_template, flash, abort, redirect, session, url_for, request, g, jsonify, make_response, Response
from app import theapp, db, dao
#from flask.ext.mail import Message
from .models import Department, TableInfo, Field, FieldType, UploadsLog
from .forms import UploadForm, NewTableInfoForm, NewFieldForm, NewDepartmentForm, BulkFieldsForm
from config import basedir, UPLOAD_FOLDER
from werkzeug import secure_filename
import os
import platform
from os import path
from openpyxl import load_workbook
from collections import Counter
import datetime
from sqlalchemy import exc
import time
from sqlalchemy import asc
#import json
import simplejson as json
from time import mktime
import pandas as pd
import numpy as np


##### some lambda functions we'll use later
resolve_cellvalue = lambda x: x.value
numeric_test = lambda x: isinstance(x, (float, int))
date_check = lambda x: isinstance(x, datetime.datetime)

################################################################

def requires_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not check_auth(auth.username, auth.password):
            return authenticate()
        return f(*args, **kwargs)
    return decorated

@theapp.route('/instructions', methods=['GET',])
def instructions():
    return render_template('instructions.html')


@theapp.route('/admin/departments')
@requires_auth
def admin():
    departments = Department.query
    return render_template('departments.html', departments = departments, asc = asc)

@theapp.route('/admin/departments/add', methods=['GET', 'POST'])
def add_department():
    form = NewDepartmentForm()

    if form.validate_on_submit():
        name = dao.stripinputstring(form.departmentname.data)
        department = Department(name = name)
        db.session.add(department)
        db.session.commit()
        return redirect(url_for('admin'))
    return render_template('add_department.html', form = form)

@theapp.route('/admin/departments/delete/<int:did>', methods=['GET', 'POST'])
def delete_department(did):
    # department
    department = Department.query.filter_by(id = did).first()

    for info in department.tables_info:
        dao.delete_tableinfo(info)
    dao.delete(department)
    return redirect(url_for('admin'))

@theapp.route('/admin/department/<int:did>/tables_info/add', methods=['GET', 'POST'])
def add_tableinfo(did):
    form = NewTableInfoForm()

    if form.validate_on_submit():
        # create & insert entry into TableInfo table
        descriptive_name = dao.stripinputstring(form.descriptive_name.data)
        table_name = dao.stripinputstring(form.table_name.data.lower())

        tableinfo = TableInfo(descriptive_name = descriptive_name, table_name = table_name, department_id = did, table_status = "1")
        db.session.add(tableinfo)

        # create the table itself
        table = db.Table(table_name, db.metadata, db.Column('id', db.Integer))

        try:
            table.create(db.engine)
            db.session.commit()
        except exc.SQLAlchemyError:
            db.session.rollback()

        return redirect(url_for('admin'))
    return render_template('add_tableinfo.html', form = form)

@theapp.route('/admin/department/tables_info/delete/<int:iid>', methods=['GET', 'POST'])
def delete_tableinfo(iid):

    # tableinfo object to be deleted
    table = TableInfo.query.filter_by(id = iid).first()

    for field in table.fields:
        db.session.delete(field)
    db.session.commit()

    # delete entry from TableInfo
    dao.delete(table)

    # drop table
    db.engine.execute('DROP TABLE {}'.format(table.table_name))
    db.session.commit()

    return redirect(url_for('admin'))

@theapp.route('/admin/tables_info/<int:iid>/fields')
def fields(iid):

    # create and add the field types if they don't already exist
    t = FieldType(name = "Text")
    n = FieldType(name = "Number")
    d = FieldType(name = "Date")

    addFieldTypes(t)
    addFieldTypes(n)
    addFieldTypes(d)
    #########################################################

    fields = Field.query.filter_by(tableinfo_id = iid)
    tableinfo = TableInfo.query.filter_by(id = iid)
    return render_template('fields.html', fields = fields, tableinfo = tableinfo, asc = asc)

@theapp.route('/admin/add_field/<int:iid>', methods=['GET', 'POST'])
def add_field(iid):

    form = NewFieldForm()
    tableinfo = TableInfo.query.filter_by(id = iid).first()
    if form.validate_on_submit():

        fieldname = dao.stripinputstring(form.fieldname.data)
        fieldtype = form.fieldtype.data

        dao.add_field(tableinfo, fieldname, fieldtype)
        return redirect(url_for('fields', iid = iid))
    return render_template('add_field.html', form = form, tableinfo = tableinfo)

#####################################################################
@theapp.route('/admin/bulk_add_fields/<int:iid>', methods=['GET', 'POST'])
def bulk_add_fields(iid):

    form = BulkFieldsForm()
    tableinfo = TableInfo.query.filter_by(id = iid).first()
    if form.validate_on_submit():

        data = form.upload.data
        file_name = secure_filename(data.filename)
        file_path = path.join(basedir,theapp.config['UPLOAD_FOLDER'], file_name)
        data.save(file_path)

        wb = load_workbook(file_path, data_only = True, read_only=True)
        sheet = wb.get_active_sheet()
        allrows = sheet.iter_rows() #this is a generator

        for row in allrows:
            if (row[0].value is not None) or (row[1].value is not None):

                fieldname = dao.stripinputstring(row[0].value) #field name stripped
                fieldtypetext = dao.stripinputstring(row[1].value)

                if fieldtypetext == "Text":
                    fieldtype = FieldType.query.filter_by(name = "Text").first()
                elif fieldtypetext == "Number":
                    fieldtype = FieldType.query.filter_by(name = "Number").first()
                elif fieldtypetext == "Date":
                    fieldtype = FieldType.query.filter_by(name = "Date").first()
                else:
                    fieldtype = FieldType.query.filter_by(name = "Text").first()

                dao.add_field(tableinfo, fieldname, fieldtype)

        return redirect(url_for('fields', iid = iid))
    return render_template('bulk_add_fields.html', form = form, tableinfo = tableinfo)

#####################################################################
@theapp.route('/admin/delete_field/<int:iid>/<int:fid>', methods=['GET', 'POST'])
def delete_field(iid, fid):

    field = Field.query.filter_by(id = fid).first()
    tableinfo = TableInfo.query.filter_by(id = iid).first()

    dao.delete_field(field,tableinfo)
    return redirect(url_for('fields', iid = iid))
####################################################################
@theapp.route('/upload/summary/<table_name>/<username>')
def summary(table_name,username):

    return render_template('summary.html', table_name = table_name, username = username)
#####################################################################
@theapp.route('/upload/preview', methods=['GET', 'POST'])
def preview():

    tablename = request.values.get("tablename")
    excel_columns = request.values.get("excel_columns")
    username = request.values.get("username")
    original_file_name = request.values.get("original_file_name")

    df = pd.read_sql(sql = "select * from {} order by id limit 10".format(tablename), con=db.engine)
    df = df.drop('id', axis = 1)

    return render_template('preview.html', tablename = tablename, data = df.to_html(classes="table table-striped", index=False, na_rep=""), username = username, original_file_name = original_file_name)


#######################################################################
@theapp.route('/upload', methods=['GET', 'POST'])
def upload():
    form = UploadForm()
    if form.validate_on_submit():

        data = form.upload.data
        original_file_name = data.filename
        file_name = secure_filename(data.filename)
        file_path = path.join(basedir,theapp.config['UPLOAD_FOLDER'], file_name)
        data.save(file_path)

        wb = load_workbook(file_path, data_only = True, read_only=True)
        sheet = wb.worksheets[0] # data should be in the first worksheet
        allrows = sheet.iter_rows() #this is a generator
        excel_columns_object = next(allrows) # row object for excel columns

        # Column names from excel file
        excel_columns = []
        for c in excel_columns_object:
            if c.value is not None: # If user uses delete button to delete contents of header cell, ignore it
                if isinstance(c.value, str) is False: # if not a string
                    return bad_request("Column names have to be strings without spaces. Value in cell {} is not a string".format(c.coordinate))
                if not c.value.strip(): # Empty string in header
                    return bad_request("Cell {} has line space(s). Delete the contents of the cell".format(c.coordinate))
                excel_columns.append(c.value.strip())


        duplicate_columns = [k for k,v in Counter(excel_columns).items() if v >1]
        if duplicate_columns: # A falsy. If there are duplicate elements
            return bad_request("Column(s) {} are duplicates".format(duplicate_columns))

        tableinfo = form.targettableinfo.data # tableinfo object in the databse order
        tablecolumns = [col.name for col in tableinfo.fields] #list comprehension. ha!

        # Check if the columns match
        if not set(excel_columns) == set(tablecolumns):
            foreign_columns = [x for x in excel_columns if x not in tablecolumns]
            if foreign_columns:
                return bad_request("Column(s) {} in your excel file are not defined. Please delete them".format(foreign_columns))
            missing_columns = [x for x in tablecolumns if x not in excel_columns]
            if missing_columns:
                return bad_request("Column(s) {} are required by the table but not found in your excel file.".format(missing_columns))

        table_coulumns_with_types = {col.name : col.type.name for col in tableinfo.fields}
        # for col in tableinfo.fields:
        #     table_coulumns_with_types[col.name] = col.type.name
        #print(table_coulumns_with_types)
        df = pd.DataFrame(allrows, columns = excel_columns)

        for col_name, col_series in df.iteritems():

            df[col_name] = df[col_name].apply(resolve_cellvalue) # resolve the cell values of column
            col_type = table_coulumns_with_types[col_name] # get the data type in that column

            if col_type == 'Date': # For date column
                #print(type(col_series[0]))
                if df[col_name].dtype != np.dtype('datetime64[ns]'):
                    non_date_rows = df[col_name][~df[col_name].apply(date_check)]
                    non_date_dict = {}
                    for i, j in non_date_rows.iteritems():
                        non_date_dict[i+2] = j

                    return bad_request('''Non-date cell(s) under the '{}' column. {}'''.format(col_name, ["Row {} has value '{}'.".format(key, value) for key, value in non_date_dict.items()]))

            elif col_type == 'Number': # For number column
                if df[col_name].dtype not in [np.dtype('float64'), np.dtype('int64')]:
                    non_numeric_rows = df[col_name][~df[col_name].apply(numeric_test)]
                    non_numeric_dict = {}
                    for i, j in non_numeric_rows.iteritems():
                        non_numeric_dict[i+2] = j
                    return bad_request('''Non-numeric cell(s) under the '{}' column. {}'''.format(col_name, ["Row {} has value '{}'.".format(key, value) for key, value in non_numeric_dict.items()]))

        df.index = df.index + 1

        tablename = tableinfo.table_name


        db.session.execute('TRUNCATE {} CASCADE'.format(tablename))
        db.session.commit()


        df.to_sql(name=tablename, con=db.engine, if_exists='append', index=True, index_label='id')

        # Log activity
        this_log = UploadsLog(user_name = "Adam Smith", user_cwsid = "adams", file_name = original_file_name, destination_table = tablename, timestamp = datetime.datetime.now())
        db.session.add(this_log)
        db.session.commit()

        #abort(422)
        username = "Adam Smith"

        return redirect(url_for('preview', tablename = tablename, excel_columns = excel_columns, username = username, original_file_name = original_file_name))

    return render_template('upload.html', title = 'Upload', form = form)

@theapp.errorhandler(422)
def bad_request(error):
    return render_template('bad_request.html', error = error), 422

def removeNonAscii(s): return "".join(i for i in s if ord(i)<128)


def addFieldTypes(thetype):
    if not FieldType.query.filter_by(name = thetype.name).first():
        dao.add(thetype)

class DateTimeEncoder(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, datetime.date):
            return o.strftime('%m/%d/%Y')

        return json.JSONEncoder.default(self, o)


def check_auth(username, password):
    """This function is called to check if a username /
    password combination is valid.
    """
    return username == 'admin' and password == 'admin'

def authenticate():
    """Sends a 401 response that enables basic auth"""
    return Response(
    'Could not verify your access level for that URL.\n'
    'You have to login with proper credentials', 401,
    {'WWW-Authenticate': 'Basic realm="Login Required"'})
